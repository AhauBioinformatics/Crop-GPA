import openpyxl
import json
import io
# excel表格转json文件

def excel_to_json(excel_file):
    # 加载工作薄
    book = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = book["Sheet1"]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    result = []
    heads = []
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}
        for column in range(max_column):
            # 读取第二行开始每一个数据
            k = heads[column]
            cell = sheet.cell(row + 1, column + 1)
            value = cell.value
            one_line[k] = value
        print(one_line)
        result.append(one_line)
    book.close()
    # 将json保存为文件

if __name__ == '__main__':
    excel_to_json("../file/process_file/2022091810300/result.xlsx")