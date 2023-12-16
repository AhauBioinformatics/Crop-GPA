#!/usr/bin/env python
# coding: utf-8
import os
import queue
import threading

import xlrd
from gensim.models import Doc2Vec
from gensim.models.doc2vec import TaggedDocument
from gensim.parsing.preprocessing import preprocess_string, remove_stopwords
import random
import warnings
import pandas as pd
import numpy as np
import requests

warnings.filterwarnings("ignore")
from itertools import product
from collections import Counter
from copy import deepcopy
from sklearn.decomposition import PCA
from flask import Flask, jsonify, request, make_response

from flask_cors import CORS

app = Flask(__name__, template_folder='templates')

parentUrl = "https://api.crop.aielab.net/crop/cropTask"


class DocumentDataset(object):
    def __init__(self, data: pd.DataFrame, column):
        document = data[column].apply(self.preprocess)
        self.documents = [TaggedDocument(text, [index])
                          for index, text in document.iteritems()]

    def preprocess(self, document):
        return preprocess_string(remove_stopwords(document))

    def __iter_(self):
        for document in self.documents:
            yield documents

    def tagged_documents(self, shuffle=False):
        if shuffle:
            random.shuffle(self.documents)
        return self.documents


import openpyxl


@app.route('/is_excel_exist', methods=['GET'])
def is_excel_exist():
    task_id = request.values.get('task_id')
    resp = {'code': 200, 'msg': 'success', 'data': ''}
    result = os.path.lexists(r'../file/process_file/' + task_id + '/result.xlsx')
    if result:
        return jsonify(resp)
    else:
        resp['code'] = -1
        return jsonify(resp)


import platform


@app.route("/downloadExcel", methods=['POST'])
def downloadExcel():
    # task_id = request.values.get('task_id')
    json_input = request.json
    task_id = json_input["task_id"]
    global file
    file = open("../file/process_file/" + task_id + "/result.xlsx", "rb").read()
    response = make_response(file)
    return response


@app.route('/index', methods=['GET'])
def index_test():
    return "hello crop-gpa deep learning !"


@app.route('/read_excel', methods=['GET'])
def read_excel():
    task_id = request.values.get('task_id')
    label = request.values.get('label')
    relevant = request.values.get('relevant')
    resp = {'code': 200, 'msg': 'get data success', 'list1': '', 'list2': ''}
    if task_id is None or len(task_id) < 1:
        resp['code'] = -1
        resp['msg'] = "please input task_id"
        return jsonify(resp)
    try:
        # 加载工作薄
        book = openpyxl.load_workbook("../file/process_file/" + task_id + "/result.xlsx")
        # 获取sheet页
        sheet = book["Sheet1"]
        # 行数
        max_row = sheet.max_row
        # 列数
        max_column = sheet.max_column
        print("max_row: %d, max_column: %d" % (max_row, max_column))
        # 结果，数组存储
        result = []
        result2 = []
        heads = []
        # 解析表头
        for column in range(max_column):
            # 读取的话行列是从（1，1）开始
            heads.append(sheet.cell(1, column + 1).value)
        # 遍历每一行
        for row in range(max_row):
            if row == 0:
                continue
            one_line = {}
            for column in range(max_column):
                # 读取第二行开始每一个数据
                k = heads[column]
                cell = sheet.cell(row + 1, column + 1)
                value = cell.value
                one_line[k] = value
            if row <= max_row / 2:
                result.append(one_line)
            else:
                result2.append(one_line)
        book.close()
        resp["list1"] = result
        resp["list2"] = result2
        return jsonify(resp)
    except:
        resp['code'] = -1
        resp['msg'] = "read excel error"
        return jsonify(resp)


# 创建一个队列
q = queue.Queue()


@app.route('/upload_file', methods=['POST'])
def upload_file():
    # file
    resp = {'code': 200, 'msg': 'file upload success', 'task_id': '', 'file_path': ''}
    import time
    # task id is current time
    task_id = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    upload_file_dir = '../file/upload_file'
    file = request.files.get('excel_file')
    types = request.values.get('type')

    if file is None:
        resp['code'] = -1
        resp['msg'] = "please upload file"
        return jsonify(resp)
    if types is None or len(types) < 1:
        resp['code'] = -1
        resp['msg'] = "please input Tg or Tt"
        return jsonify(resp)

    allowed_file_type = {'xls', 'xlsx'}
    allowed_file = file.filename.split('.', 1)[1].lower()
    # file name
    file_name = task_id + '.' + allowed_file
    if allowed_file in allowed_file_type:
        path = os.path.join(upload_file_dir, file_name)
        file.save(path)
        resp['task_id'] = task_id
        resp['file_path'] = path
        return jsonify(resp)
    else:
        resp['code'] = -1
        resp['msg'] = 'The file is not an Excel file !'
        return jsonify(resp)


@app.route('/main_upload', methods=["POST"])
def main_upload():
    resp = {'code': 200, 'msg': 'task is running'}

    if request.method == 'POST':
        # receive params
        param_type = request.values.get('type')
        email = request.values.get('email')
        name = request.values.get('name')
        task_id = request.values.get('task_id')
        file_path = request.values.get('file_path')
        # 判断参数是否为空以及是否合法
        if param_type is None or len(param_type) < 1:
            resp['code'] = -1
            resp['msg'] = "please input Tg or Tt"
            return jsonify(resp)
        if task_id is None or len(task_id) < 1:
            resp['code'] = -1
            resp['msg'] = "please input task_id"
            return jsonify(resp)
        if name is None or len(name) < 1:
            resp['code'] = -1
            resp['msg'] = "please input name"
            return jsonify(resp)
        if email is None or len(email) < 1:
            resp['code'] = -1
            resp['msg'] = "please input email"
            return jsonify(resp)
        if file_path is None or len(file_path) < 1:
            resp['code'] = -1
            resp['msg'] = "please input file"
            return jsonify(resp)
        # 把线程放进队列中
        if q.qsize() == 0:  # 队列里面当前没有任务在跑,每次只能跑一个子线程的任务
            # 插入数据到数据库并且任务为执行中的状态
            data = {
                'taskId': task_id,
                'name': name,
                'email': email,
                'filePath': file_path,
                'type': param_type,
                'status': 'running'
            }
            # 字符串格式
            res = requests.post(url=parentUrl + "/create", json=data,
                                headers={'Content-Type': 'application/json;charset=UTF-8'})
            print(res.text)
            items = [task_id, name, email, param_type, file_path]
            q.put(items)
            # 启动一个子线程开始跑算法
            t = threading.Thread(target=main,
                                 args=(
                                     task_id, name, email, param_type, file_path, q),
                                 name='deeplearn')
            t.start()

            return jsonify(resp)
        else:
            # 插入数据到数据库并且任务为等待中的状态
            data = {
                'taskId': task_id,
                'name': name,
                'email': email,
                'filePath': file_path,
                'type': param_type,
                'status': 'waiting'
            }
            # 字符串格式
            res = requests.post(url=parentUrl + "/create", json=data,
                                headers={'Content-Type': 'application/json;charset=UTF-8'})
            print(res.text)
            resp['msg'] = 'The previous task is running and your task is being queued!'
            return jsonify(resp)


import two_step

import crop_gpa_gcn

# 执行的主函数
def main(task_id, name, email, param_type, file_path, q: queue.Queue):
    print("main task is running")
    print(task_id)
    try:
        select = param_type
        one = os.path.lexists(r'../file/process_file/' + task_id + '/task_Tg__testlabel0_knn5neighbors_edge.npz')
        two = os.path.lexists(r'../file/process_file/' + task_id + '/node_feature_label.csv')
        three = os.path.lexists(
            r'../file/process_file/' + task_id + '/task_Tg__testlabel0_knn_edge_train_test_index_all.npz')
        adj_val = os.path.lexists(r'../file/process_file/' + task_id + '/adj_val.npz')

        # 第二步文件有了,直接执行第三步，执行完了弹栈
        if adj_val:
            print(task_id + " two data exist, so run three_step")
            train_saveys.train_saveys_main(task_id, email, name)
            print(task_id + " three_step end")
        elif one and two and three:  # 如果第一步文件有了 就直接跑第二步
            print(task_id + " one data exist, so run two_step")
            two_step.two_step_run(task_id)
            print(task_id + " two_step ended")
            train_saveys.train_saveys_main(task_id, email, name)
        else:  # 前面的文件都没有，就按照顺序一步步的来执行
            # 传入文件的地址
            one_step(task_id, select, file_path)
            print(task_id + " one step end")

            two_step.two_step_run(task_id, select)
            print(task_id + "two step end")

            train_saveys.train_saveys_main(task_id, email, name)
            print(task_id + " three_step end")

        print(task_id + "send email success")

        # 弹栈
        q.get()
        # 更新这个任务在数据库中的执行状态，通过掉接口的方式更新
        update_task_status(task_id, "completed")
        # 判断当前数据库中是否存在wait的任务，如果有，获取信息任务的基本信息，然后执行
        if q.qsize() == 0:
            ret = requests.get(parentUrl + "/getTask")
            result = ret.json()
            data = result.get("data")
            if result.get("code") == 200:
                items = [str(data.get("taskId")), str(data.get("name")), str(data.get("email")), str(data.get("type")),
                         str(data.get("filePath"))]
                q.put(items)
                # 启动一个子线程开始跑算法
                t = threading.Thread(target=main,
                                     args=(
                                         str(data.get("taskId")), str(data.get("name")), str(data.get("email")),
                                         str(data.get("type")), str(data.get("filePath")), q),
                                     name='deeplearn')
                t.start()
                update_task_status(str(data.get("taskId")), "running")
    except:
        # 报错
        print(task_id + " task execute error")
        # 弹栈
        q.get()
        # send email
        import to_email
        subject = "Prediction Error Reminder"
        message = "Your task " + task_id + " Error, please contact platform administrator: aielab_crop@163.com"
        to_email.send_email_file(email, message, None, subject, name, task_id)
        # update task status
        update_task_status(task_id, "error")
        # execute next task
        executeNextTask()


# 更新任务状态
def update_task_status(task_id, status):
    data = {
        'status': status
    }
    url = parentUrl + "/update/?taskId=" + task_id
    # 字符串格式
    requests.post(url=url, json=data, headers={'Content-Type': 'application/json;charset=UTF-8'})
    return


# 手动接口，如果running任务报错了，可以手动临时把任务重新启动起来
@app.route('/execute_run_task', methods=["GET"])
def execute_run_task():
    resp = {'code': 200, 'msg': 'task is running'}

    task_id = request.values.get('task_id')
    # 判断参数是否为空以及是否合法
    if task_id is None or len(task_id) < 1:
        resp['code'] = -1
        resp['msg'] = "please input task_id"
        return jsonify(resp)

    if q.qsize() == 0:
        param = {
            "taskId": task_id
        }
        ret = requests.get(parentUrl + "/getOneTask", params=param)
        result = ret.json()
        data = result.get("data")
        print(result)
        if result.get("code") == 200:
            items = [str(data.get("taskId")), str(data.get("name")), str(data.get("email")), str(data.get("type")),
                     str(data.get("filePath"))]
            q.put(items)
            # 启动一个子线程开始跑算法
            t = threading.Thread(target=main,
                                 args=(
                                     str(data.get("taskId")), str(data.get("name")), str(data.get("email")),
                                     str(data.get("type")), str(data.get("filePath")), q),
                                 name='deeplearn')
            t.start()
        update_task_status(task_id, "running")
        return jsonify(resp)
    else:
        resp["msg"] = "query full"
        return jsonify(resp)


@app.route('/execute_wait_task', methods=["GET"])
def executeNextTask():
    resp = {'code': 200, 'msg': 'task is running'}
    if q.qsize() == 0:
        ret = requests.get(parentUrl + "/getTask")
        result = ret.json()
        data = result.get("data")
        if result.get("code") == 200:
            items = [str(data.get("taskId")), str(data.get("name")), str(data.get("email")), str(data.get("type")),
                     str(data.get("filePath"))]
            q.put(items)
            # 启动一个子线程开始跑算法
            t = threading.Thread(target=main,
                                 args=(
                                     str(data.get("taskId")), str(data.get("name")), str(data.get("email")),
                                     str(data.get("type")), str(data.get("filePath")), q),
                                 name='deeplearn')
            t.start()
        return jsonify(resp)
    else:
        resp['msg'] = "please wait top task end"
        return jsonify(resp)


def one_step(task_id, select, file_path):
    text = pd.read_excel(file_path)
    head = list(text)
    name_list = text[head[0]].tolist()

    docVecModel = Doc2Vec(min_count=1,
                          window=5,
                          vector_size=100, sample=1e-4,
                          negative=5,
                          workers=2)

    path = r'../file/process_file/' + task_id
    if not os.path.exists(path):
        os.mkdir(path)

    if select == 'Tt':
        new_col = ['trait', 'trait_def']
        text.columns = new_col
        trait_2 = pd.read_excel(r'../data/peco_def.xlsx')[['trait', 'trait_def']]
        trait = pd.concat([text, trait_2]).drop_duplicates(['trait']).reset_index(drop=True)
        trait['trait'].to_excel(path + '/trait_name.xlsx')
        document_dataset = DocumentDataset(trait, 'trait_def')
        # TaggedDocument的实例
        docVecModel.build_vocab(document_dataset.tagged_documents())
        docVecModel.train(document_dataset.tagged_documents(shuffle=True),
                          total_examples=docVecModel.corpus_count,
                          epochs=10)
        pl = []
        for i in range(len(trait['trait'])):
            pl.append(docVecModel[i])
        # 皮尔逊相关系数
        TSSM = np.corrcoef(pl)
    else:
        trait = pd.read_excel(r'../data/peco_def.xlsx')[['trait', 'trait_def']]
        trait['trait'].to_excel(path + '/trait_name.xlsx')
        TSSM = np.loadtxt(r'../data/PSSM.txt', dtype=np.float32, delimiter='\t')

    # 基因相似矩阵

    if select == 'Tg':
        new_col = ['gene', 'gene_def']
        text.columns = new_col
        gene_2 = pd.read_excel(r'../data/gene_name_def.xlsx')[['gene', 'gene_def']]
        gene = pd.concat([text, gene_2]).drop_duplicates(['gene']).reset_index(drop=True)
        gene['gene'].to_excel(path + '/gene_name.xlsx')
        document_dataset = DocumentDataset(gene, 'gene_def')
        # TaggedDocument的实例
        docVecModel.build_vocab(document_dataset.tagged_documents())
        docVecModel.train(document_dataset.tagged_documents(shuffle=True),
                          total_examples=docVecModel.corpus_count,
                          epochs=10)
        gl = []
        for i in range(len(gene['gene'])):
            gl.append(docVecModel[i])

        GSSM = np.corrcoef(gl)

        pca = PCA(n_components=10)  # 把原数据降成10维
        GSSM = pca.fit_transform(GSSM)

    else:
        gene = pd.read_excel(r'../data/gene_name_def.xlsx')[['gene', 'gene_def']]
        gene['gene'].to_excel(path + '/gene_name.xlsx')
        GSSM = np.loadtxt(r'../data/GSSM_.txt', dtype=np.float32)

    test = []
    for i in name_list:
        if select == 'Tt':
            test.append(int(trait.loc[trait['trait'] == i].index.values))
        else:
            test.append(int(gene.loc[gene['gene'] == i].index.values))

    trait['trait_idx'] = trait.index
    gene['gene_idx'] = gene.index

    # 已知关系对
    known_associations = pd.read_excel(r'../data/known_associations.xlsx', header=None, names=['trait', 'gene'])
    known_associations = pd.merge(known_associations, trait[['trait', 'trait_idx']], on='trait')
    known_associations = pd.merge(known_associations, gene[['gene', 'gene_idx']], on='gene')
    known_associations['label'] = 1
    known_associations = known_associations.drop_duplicates().reset_index(drop=True)

    # 正样本

    positive_gtp = known_associations[['trait_idx', 'gene_idx']].sort_values(by=['trait_idx', 'gene_idx'],
                                                                             ascending=True).reset_index(drop=True)
    peco_idx = list(set(trait['trait_idx']))
    gene_idx = list(set(gene['gene_idx']))
    # 所有样本
    all_gtp = pd.DataFrame(list(product(peco_idx, gene_idx)), columns=['trait_idx', 'gene_idx'])
    # 负样本
    negative_gtp = pd.concat([all_gtp, positive_gtp], axis=0).drop_duplicates(keep=False)
    positive_gtp['label'] = 1
    negative_gtp['label'] = 0
    all_gtp = pd.concat([positive_gtp, negative_gtp], axis=0)

    if select == 'Tg':
        item = 'gene_idx'
        IPE, IG, dtp, gene_ids, trait_ids, knn_x, label = obtain_data(TSSM, GSSM, test, item, all_gtp)
        ids = gene_ids
        train_index_all, test_index_all, train_id_all, test_id_all = generate_task_Tg_Tt_train_test_idx(item, ids,
                                                                                                        test,
                                                                                                        dtp)
    else:
        item = 'trait_idx'
        IPE, IG, dtp, gene_ids, trait_ids, knn_x, label = obtain_data(TSSM, GSSM, test, item, all_gtp)
        ids = trait_ids
        train_index_all, test_index_all, train_id_all, test_id_all = generate_task_Tg_Tt_train_test_idx(item, ids,
                                                                                                        test,
                                                                                                        dtp)


    np.savez_compressed(path + '/task_' + select + '__testlabel0_knn_edge_train_test_index_all.npz',
                        train_index_all=train_index_all,
                        test_index_all=test_index_all,
                        train_id_all=train_id_all,
                        test_id_all=test_id_all)
    pwd = (path + '/')
    knn_x, knn_y, knn, knn_neighbors_graph = generate_knn_graph_save(knn_x, label, train_index_all, test_index_all,
                                                                     pwd,
                                                                     select)
    node_feature_label = pd.concat([dtp, knn_x], axis=1).fillna(0)
    # 存一下这个文件
    node_feature_label.to_csv(pwd + 'node_feature_label.csv')


# text_associations
def load_data(TSSM, GSSM):
    IT = pd.DataFrame(TSSM).reset_index()
    IG = pd.DataFrame(GSSM).reset_index()
    IT.rename(columns={'index': 'id'}, inplace=True)
    IG.rename(columns={'index': 'id'}, inplace=True)
    IT['id'] = IT['id']
    IG['id'] = IG['id']

    return IT, IG


def sample(all_gtp, random_seed, test, item):
    known_associations = all_gtp.loc[all_gtp['label'] == 1]
    unknown_associations = all_gtp.loc[all_gtp['label'] == 0]

    random_negative = unknown_associations.sample(n=known_associations.shape[0], random_state=random_seed, axis=0)
    # print(len(random_negative))
    test_associations = all_gtp[all_gtp[item].isin(test)]
    # print(len(test_associations))
    sample_df = known_associations.append(random_negative)
    # print(len(sampl_df))
    sample_df = sample_df.append(test_associations).drop_duplicates()
    # print(len(sample_df))
    sample_df.reset_index(drop=True, inplace=True)

    return sample_df


def obtain_data(TSSM, GSSM, test, item, all_gtp):
    IT, IG = load_data(TSSM, GSSM)

    dtp = sample(all_gtp, 1234, test, item)

    gene_ids = list(set(dtp['gene_idx']))
    trait_ids = list(set(dtp['trait_idx']))
    random.shuffle(gene_ids)
    random.shuffle(trait_ids)

    knn_x = pd.merge(dtp, IT, left_on='trait_idx', right_on='id')
    knn_x = pd.merge(knn_x, IG, left_on='gene_idx', right_on='id')

    label = dtp['label'].fillna(0)
    knn_x.drop(labels=['gene_idx', 'trait_idx', 'label', 'id_x', 'id_y'], axis=1, inplace=True)
    print(IT.shape[0] + IG.shape[0])
    print(knn_x.shape[1])
    # assert IPE.shape[0] + IG.shape[0] == knn_x.shape[1]
    print(knn_x.shape, Counter(label))

    return IT, IG, dtp, gene_ids, trait_ids, knn_x, label


def generate_task_Tg_Tt_train_test_idx(item, ids, test, dtp):
    train_index_all, test_index_all = [], []
    train_id_all, test_id_all = [], []

    train_ids = list(set(ids) ^ set(test))

    test_idx = dtp[dtp[item].isin(list(test))].index.tolist()

    train_idx = dtp[dtp[item].isin(train_ids)].index.tolist()
    random.shuffle(test_idx)
    random.shuffle(train_idx)
    print('# Pairs: Train = {} | Test = {}'.format(len(train_idx), len(test_idx)))
    assert len(train_idx) + len(test_idx) == len(dtp)

    train_index_all.append(train_idx)
    test_index_all.append(test_idx)

    train_id_all.append(train_ids)
    test_id_all.append(test)

    print('train_index_all', train_index_all)
    print('test_index_all', test_index_all)
    print('train_id_all', train_id_all)
    print('test_id_all', test_id_all)

    return train_index_all, test_index_all, train_id_all, test_id_all


# knn进行图的生成
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import confusion_matrix
from sklearn.metrics import roc_auc_score, auc
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import classification_report
import scipy.sparse as sp


def generate_knn_graph_save(knn_x, label, train_index_all, test_index_all, pwd, task):
    knn_y = deepcopy(label)

    print('Label: ', Counter(label))
    print('knn_y: ', Counter(knn_y))

    knn = KNeighborsClassifier(n_neighbors=5)
    knn.fit(knn_x, knn_y)

    knn_y_pred = knn.predict(knn_x)
    knn_y_prob = knn.predict_proba(knn_x)
    # knn.kneighbors_graph计算相邻节点之间的权重
    knn_neighbors_graph = knn.kneighbors_graph(knn_x, n_neighbors=5)

    # classification_report模型评估报告
    prec_reca_f1_supp_report = classification_report(knn_y, knn_y_pred, target_names=['label_0', 'label_1'])
    tn, fp, fn, tp = confusion_matrix(knn_y, knn_y_pred).ravel()

    pos_acc = tp / sum(knn_y)
    neg_acc = tn / (len(knn_y_pred) - sum(knn_y_pred))  # [y_true=0 & y_pred=0] / y_pred=0
    accuracy = (tp + tn) / (tn + fp + fn + tp)

    recall = tp / (tp + fn)
    precision = tp / (tp + fp)
    f1 = 2 * precision * recall / (precision + recall)

    roc_auc = roc_auc_score(knn_y, knn_y_prob[:, 1])
    prec, reca, _ = precision_recall_curve(knn_y, knn_y_prob[:, 1])
    aupr = auc(reca, prec)

    print(
        'acc={:.4f}|precision={:.4f}|recall={:.4f}|f1={:.4f}|auc={:.4f}|aupr={:.4f}|pos_acc={:.4f}|neg_acc={:.4f}'.format(
            accuracy, precision, recall, f1, roc_auc, aupr, pos_acc, neg_acc))
    print('tn = {}, fp = {}, fn = {}, tp = {}'.format(tn, fp, fn, tp))
    print('y_pred: ', Counter(knn_y_pred))
    print('y_true: ', Counter(knn_y))
    #       print('knn_score = {:.4f}'.format(knn.score(knn_x, knn_y)))

    # 存一下这个文件
    sp.save_npz(pwd + 'task_' + task + '' + '__testlabel0_knn' + str(5) + 'neighbors_edge' + '.npz',
                knn_neighbors_graph)

    return knn_x, knn_y, knn, knn_neighbors_graph


CORS(app, resources=r'/*')
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=2100)
    # text = pd.read_excel("../file/upload_file\\20220918183002.xlsx")
    # print(list(text))
